# -*- coding: utf-8 -*-
"""
Excel 輸出模組
Excel Output Module

生成 Excel 格式的分析報告和數據導出
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import os

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl 未安裝，Excel 功能將受限")


class ExcelExporter:
    """Excel 導出器"""

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("需要安裝 openpyxl 套件: pip install openpyxl")

        self.workbook = None
        self.styles = self._define_styles()

    def _define_styles(self) -> Dict[str, Any]:
        """定義 Excel 樣式"""
        return {
            'header': Font(bold=True, size=12, color="FFFFFF"),
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'subheader': Font(bold=True, size=11, color="000000"),
            'subheader_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            'data': Font(size=10),
            'number': Font(size=10),
            'positive': Font(size=10, color="008000"),
            'negative': Font(size=10, color="FF0000"),
            'warning': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'success': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }

    def create_comprehensive_report(self, analysis_data: Dict[str, Any], filename: str = None) -> str:
        """創建綜合分析報告"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"股票分析報告_{timestamp}.xlsx"

            self.workbook = openpyxl.Workbook()

            # 刪除默認工作表
            self.workbook.remove(self.workbook.active)

            # 創建各個工作表
            self._create_summary_sheet(analysis_data)
            self._create_technical_sheet(analysis_data)
            self._create_strategy_sheet(analysis_data)
            self._create_performance_sheet(analysis_data)

            # 保存文件
            self.workbook.save(filename)
            logger.info(f"Excel 報告已保存至: {filename}")

            return filename

        except Exception as e:
            logger.error(f"創建 Excel 報告失敗: {e}")
            return ""

    def _create_summary_sheet(self, data: Dict[str, Any]):
        """創建總結工作表"""
        sheet = self.workbook.create_sheet("總結報告")

        # 標題
        sheet['A1'] = "股票分析綜合報告"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:E1')

        # 基本信息
        row = 3
        sheet.cell(row, 1, "股票代碼").font = self.styles['subheader']
        sheet.cell(row, 1).fill = self.styles['subheader_fill']
        sheet.cell(row, 2, data.get('stock_code', 'N/A'))

        row += 1
        sheet.cell(row, 1, "分析日期").font = self.styles['subheader']
        sheet.cell(row, 1).fill = self.styles['subheader_fill']
        sheet.cell(row, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # 關鍵指標
        row += 2
        sheet.cell(row, 1, "關鍵指標").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:E{row}')

        headers = ["指標名稱", "數值", "評價", "建議"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row + 1, col, header).font = self.styles['subheader']
            sheet.cell(row + 1, col).fill = self.styles['subheader_fill']

        # 技術指標數據
        tech_data = data.get('technical_data', {})
        summary_data = [
            ["當前價格", tech_data.get('current_price', 0), self._evaluate_price(tech_data), ""],
            ["RSI", tech_data.get('rsi', 50), self._evaluate_rsi(tech_data.get('rsi', 50)), ""],
            ["趨勢", data.get('trend', {}).get('trend', '未知'), self._evaluate_trend(data.get('trend', {})), ""],
            ["位階", data.get('position', {}).get('stage', '未知'), self._evaluate_position(data.get('position', {})), ""],
            ["綜合評分", data.get('overall_score', {}).get('總分', 0), self._evaluate_score(data.get('overall_score', {})), ""]
        ]

        for i, (name, value, evaluation, suggestion) in enumerate(summary_data):
            sheet.cell(row + 2 + i, 1, name)
            sheet.cell(row + 2 + i, 2, value)
            sheet.cell(row + 2 + i, 3, evaluation)
            sheet.cell(row + 2 + i, 4, suggestion)

        # 投資建議
        row += len(summary_data) + 3
        sheet.cell(row, 1, "投資建議").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:E{row}')

        recommendation = data.get('investment_recommendation', {})
        suggestions = [
            f"操作建議: {recommendation.get('操作建議', '未知')}",
            f"倉位建議: {recommendation.get('倉位建議', '未知')}",
            f"持有期間: {recommendation.get('持有期間', '未知')}"
        ]

        for i, suggestion in enumerate(suggestions):
            sheet.cell(row + 1 + i, 1, suggestion)
            sheet.merge_cells(f'A{row + 1 + i}:E{row + 1 + i}')

        # 應用樣式
        self._apply_sheet_styles(sheet)

    def _create_technical_sheet(self, data: Dict[str, Any]):
        """創建技術分析工作表"""
        sheet = self.workbook.create_sheet("技術分析")

        # 標題
        sheet['A1'] = "技術指標詳情"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:F1')

        # 均線系統
        row = 3
        sheet.cell(row, 1, "均線系統").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:F{row}')

        ma_data = data.get('technical_data', {})
        ma_headers = ["指標", "數值", "相對位置", "趨勢"]
        for col, header in enumerate(ma_headers, 1):
            sheet.cell(row + 1, col, header).font = self.styles['subheader']
            sheet.cell(row + 1, col).fill = self.styles['subheader_fill']

        ma_info = [
            ["5日均線", ma_data.get('sma_5', 0), self._calculate_ma_position(ma_data, 'sma_5'), "短期趨勢"],
            ["20日均線", ma_data.get('sma_20', 0), self._calculate_ma_position(ma_data, 'sma_20'), "中期趨勢"],
            ["60日均線", ma_data.get('sma_60', 0), self._calculate_ma_position(ma_data, 'sma_60'), "長期趨勢"]
        ]

        for i, (name, value, position, trend) in enumerate(ma_info):
            sheet.cell(row + 2 + i, 1, name)
            sheet.cell(row + 2 + i, 2, value)
            sheet.cell(row + 2 + i, 3, position)
            sheet.cell(row + 2 + i, 4, trend)

        # 技術指標
        row += len(ma_info) + 3
        sheet.cell(row, 1, "技術指標").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:F{row}')

        tech_headers = ["指標名稱", "數值", "狀態", "意義"]
        for col, header in enumerate(tech_headers, 1):
            sheet.cell(row + 1, col, header).font = self.styles['subheader']
            sheet.cell(row + 1, col).fill = self.styles['subheader_fill']

        tech_indicators = [
            ["RSI", ma_data.get('rsi', 50), self._get_rsi_status(ma_data.get('rsi', 50)), "動能指標"],
            ["MACD", ma_data.get('macd', 0), self._get_macd_status(data.get('trend', {})), "趨勢指標"],
            ["ATR", ma_data.get('atr', 0), "波動率", "風險指標"],
            ["布林通道", data.get('trend', {}).get('bollinger_status', '未知'), "位置", "支撐阻力"]
        ]

        for i, (name, value, status, meaning) in enumerate(tech_indicators):
            sheet.cell(row + 2 + i, 1, name)
            sheet.cell(row + 2 + i, 2, value)
            sheet.cell(row + 2 + i, 3, status)
            sheet.cell(row + 2 + i, 4, meaning)

        # 關鍵價位
        row += len(tech_indicators) + 3
        sheet.cell(row, 1, "關鍵價位").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:F{row}')

        levels = data.get('levels', {})
        level_data = [
            ["支撐位", levels.get('support', 0), "買入參考", ""],
            ["阻力位", levels.get('resistance', 0), "賣出參考", ""],
            ["買入區", levels.get('buy_zone', 0), "積極買入", ""],
            ["賣出區", levels.get('sell_zone', 0), "考慮賣出", ""],
            ["停損價", levels.get('stop_loss_buy', 0), "風險控制", ""],
            ["目標價", levels.get('target_price', 0), "利潤目標", ""]
        ]

        for i, (name, value, reference, note) in enumerate(level_data):
            sheet.cell(row + 1 + i, 1, name)
            sheet.cell(row + 2 + i, 2, value)
            sheet.cell(row + 2 + i, 3, reference)
            sheet.cell(row + 2 + i, 4, note)

        self._apply_sheet_styles(sheet)

    def _create_strategy_sheet(self, data: Dict[str, Any]):
        """創建策略建議工作表"""
        sheet = self.workbook.create_sheet("策略建議")

        # 標題
        sheet['A1'] = "投資策略建議"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:D1')

        # 策略建議
        row = 3
        strategy = data.get('strategy', {})

        strategies = [
            ("短期策略 (1-5天)", strategy.get('short_term', '無建議')),
            ("中期策略 (1-3個月)", strategy.get('medium_term', '無建議')),
            ("長期策略 (3個月以上)", strategy.get('long_term', '無建議'))
        ]

        for i, (period, suggestion) in enumerate(strategies):
            sheet.cell(row + i, 1, period).font = self.styles['subheader']
            sheet.cell(row + i, 1).fill = self.styles['subheader_fill']
            sheet.cell(row + i, 2, suggestion)
            sheet.merge_cells(f'B{row + i}:D{row + i}')

        # 風險評估
        row += len(strategies) + 2
        sheet.cell(row, 1, "風險評估").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:D{row}')

        risk = data.get('risk_assessment', {})
        risk_items = [
            ("風險等級", risk.get('風險等級', '未知')),
            ("風險評估", risk.get('風險評估', '未知'))
        ]

        for i, (item, value) in enumerate(risk_items):
            sheet.cell(row + 1 + i, 1, item)
            sheet.cell(row + 1 + i, 2, value)

        # 風險因子
        row += len(risk_items) + 2
        sheet.cell(row, 1, "風險因子").font = self.styles['subheader']
        sheet.cell(row, 1).fill = self.styles['subheader_fill']

        risk_factors = risk.get('風險因子', [])
        for i, factor in enumerate(risk_factors):
            sheet.cell(row + 1 + i, 1, f"• {factor}")

        self._apply_sheet_styles(sheet)

    def _create_performance_sheet(self, data: Dict[str, Any]):
        """創建績效統計工作表"""
        sheet = self.workbook.create_sheet("績效統計")

        # 標題
        sheet['A1'] = "績效統計數據"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:C1')

        # 績效指標
        row = 3
        sheet.cell(row, 1, "績效指標").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:C{row}')

        performance = data.get('overall_score', {})
        perf_data = [
            ("總分", performance.get('總分', 0)),
            ("等級", performance.get('等級', 'F')),
            ("評價", performance.get('評價', '未知'))
        ]

        for i, (metric, value) in enumerate(perf_data):
            sheet.cell(row + 1 + i, 1, metric)
            sheet.cell(row + 1 + i, 2, value)

        # 評分細項
        row += len(perf_data) + 2
        sheet.cell(row, 1, "評分細項").font = self.styles['subheader']
        sheet.cell(row, 1).fill = self.styles['subheader_fill']
        sheet.merge_cells(f'A{row}:C{row}')

        score_details = performance.get('評分細項', {})
        detail_data = [
            ("收益率評分", score_details.get('收益率評分', 0)),
            ("夏普比率評分", score_details.get('夏普比率評分', 0)),
            ("回撤評分", score_details.get('回撤評分', 0)),
            ("勝率評分", score_details.get('勝率評分', 0))
        ]

        for i, (item, score) in enumerate(detail_data):
            sheet.cell(row + 1 + i, 1, item)
            sheet.cell(row + 1 + i, 2, score)

        self._apply_sheet_styles(sheet)

    def export_portfolio_data(self, portfolio_data: Dict[str, Any], filename: str = None) -> str:
        """導出投資組合數據"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"投資組合數據_{timestamp}.xlsx"

            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)

            # 創建工作表
            self._create_portfolio_sheet(portfolio_data)
            self._create_trades_sheet(portfolio_data)

            self.workbook.save(filename)
            logger.info(f"投資組合數據已導出至: {filename}")

            return filename

        except Exception as e:
            logger.error(f"導出投資組合數據失敗: {e}")
            return ""

    def _create_portfolio_sheet(self, data: Dict[str, Any]):
        """創建投資組合工作表"""
        sheet = self.workbook.create_sheet("投資組合")

        # 標題
        sheet['A1'] = "投資組合持倉"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:G1')

        # 表頭
        headers = ["股票代碼", "股票名稱", "持倉數量", "平均成本", "當前價格", "市值", "盈虧"]
        row = 3
        for col, header in enumerate(headers, 1):
            sheet.cell(row, col, header).font = self.styles['subheader']
            sheet.cell(row, col).fill = self.styles['subheader_fill']

        # 持倉數據
        holdings = data.get('holdings', [])
        for i, holding in enumerate(holdings):
            sheet.cell(row + 1 + i, 1, holding.get('code', ''))
            sheet.cell(row + 1 + i, 2, holding.get('name', ''))
            sheet.cell(row + 1 + i, 3, holding.get('quantity', 0))
            sheet.cell(row + 1 + i, 4, holding.get('avg_cost', 0))
            sheet.cell(row + 1 + i, 5, holding.get('current_price', 0))
            sheet.cell(row + 1 + i, 6, holding.get('market_value', 0))
            sheet.cell(row + 1 + i, 7, holding.get('profit_loss', 0))

            # 盈虧顏色
            profit_cell = sheet.cell(row + 1 + i, 7)
            if holding.get('profit_loss', 0) > 0:
                profit_cell.font = self.styles['positive']
            elif holding.get('profit_loss', 0) < 0:
                profit_cell.font = self.styles['negative']

        self._apply_sheet_styles(sheet)

    def _create_trades_sheet(self, data: Dict[str, Any]):
        """創建交易記錄工作表"""
        sheet = self.workbook.create_sheet("交易記錄")

        # 標題
        sheet['A1'] = "交易記錄明細"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:H1')

        # 表頭
        headers = ["日期", "股票代碼", "動作", "數量", "價格", "金額", "手續費", "備註"]
        row = 3
        for col, header in enumerate(headers, 1):
            sheet.cell(row, col, header).font = self.styles['subheader']
            sheet.cell(row, col).fill = self.styles['subheader_fill']

        # 交易數據
        trades = data.get('trades', [])
        for i, trade in enumerate(trades):
            sheet.cell(row + 1 + i, 1, trade.get('date', ''))
            sheet.cell(row + 1 + i, 2, trade.get('stock_code', ''))
            sheet.cell(row + 1 + i, 3, trade.get('action', ''))
            sheet.cell(row + 1 + i, 4, trade.get('quantity', 0))
            sheet.cell(row + 1 + i, 5, trade.get('price', 0))
            sheet.cell(row + 1 + i, 6, trade.get('amount', 0))
            sheet.cell(row + 1 + i, 7, trade.get('fee', 0))
            sheet.cell(row + 1 + i, 8, trade.get('note', ''))

        self._apply_sheet_styles(sheet)

    def export_strategy_backtest(self, backtest_data: Dict[str, Any], filename: str = None) -> str:
        """導出策略回測結果"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"策略回測結果_{timestamp}.xlsx"

            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)

            # 創建工作表
            self._create_backtest_summary_sheet(backtest_data)
            self._create_backtest_trades_sheet(backtest_data)
            self._create_backtest_metrics_sheet(backtest_data)

            self.workbook.save(filename)
            logger.info(f"策略回測結果已導出至: {filename}")

            return filename

        except Exception as e:
            logger.error(f"導出策略回測結果失敗: {e}")
            return ""

    def _create_backtest_summary_sheet(self, data: Dict[str, Any]):
        """創建回測總結工作表"""
        sheet = self.workbook.create_sheet("回測總結")

        # 基本信息
        sheet['A1'] = "策略回測報告"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:D1')

        row = 3
        info_data = [
            ("策略名稱", data.get('strategy_name', '未知')),
            ("回測期間", data.get('backtest_period', '未知')),
            ("初始資金", data.get('initial_capital', 0)),
            ("最終價值", data.get('final_value', 0)),
            ("總收益率", data.get('total_return', 0)),
            ("年化收益率", data.get('annual_return', 0))
        ]

        for i, (label, value) in enumerate(info_data):
            sheet.cell(row + i, 1, label).font = self.styles['subheader']
            sheet.cell(row + i, 1).fill = self.styles['subheader_fill']
            sheet.cell(row + i, 2, value)

        # 風險指標
        row += len(info_data) + 2
        sheet.cell(row, 1, "風險指標").font = self.styles['header']
        sheet.cell(row, 1).fill = self.styles['header_fill']
        sheet.merge_cells(f'A{row}:D{row}')

        risk_data = [
            ("夏普比率", data.get('sharpe_ratio', 0)),
            ("最大回撤", data.get('max_drawdown', 0)),
            ("勝率", data.get('win_rate', 0)),
            ("總交易次數", data.get('total_trades', 0))
        ]

        for i, (label, value) in enumerate(risk_data):
            sheet.cell(row + 1 + i, 1, label)
            sheet.cell(row + 1 + i, 2, value)

        self._apply_sheet_styles(sheet)

    def _create_backtest_trades_sheet(self, data: Dict[str, Any]):
        """創建回測交易工作表"""
        sheet = self.workbook.create_sheet("回測交易")

        # 表頭
        headers = ["日期", "動作", "股票代碼", "價格", "數量", "金額", "盈虧", "累計收益"]
        row = 1
        for col, header in enumerate(headers, 1):
            sheet.cell(row, col, header).font = self.styles['subheader']
            sheet.cell(row, col).fill = self.styles['subheader_fill']

        # 交易數據
        trades = data.get('trade_details', [])
        for i, trade in enumerate(trades):
            sheet.cell(row + 1 + i, 1, trade.get('date', ''))
            sheet.cell(row + 1 + i, 2, trade.get('action', ''))
            sheet.cell(row + 1 + i, 3, trade.get('stock_code', ''))
            sheet.cell(row + 1 + i, 4, trade.get('price', 0))
            sheet.cell(row + 1 + i, 5, trade.get('quantity', 0))
            sheet.cell(row + 1 + i, 6, trade.get('amount', 0))
            sheet.cell(row + 1 + i, 7, trade.get('profit_loss', 0))
            sheet.cell(row + 1 + i, 8, trade.get('cumulative_return', 0))

        self._apply_sheet_styles(sheet)

    def _create_backtest_metrics_sheet(self, data: Dict[str, Any]):
        """創建回測指標工作表"""
        sheet = self.workbook.create_sheet("詳細指標")

        # 月度收益
        sheet['A1'] = "月度收益統計"
        sheet['A1'].font = self.styles['header']
        sheet['A1'].fill = self.styles['header_fill']
        sheet.merge_cells('A1:D1')

        monthly_data = data.get('monthly_returns', [])
        headers = ["月份", "收益", "交易次數", "勝率"]
        row = 3
        for col, header in enumerate(headers, 1):
            sheet.cell(row, col, header).font = self.styles['subheader']
            sheet.cell(row, col).fill = self.styles['subheader_fill']

        for i, month_data in enumerate(monthly_data):
            sheet.cell(row + 1 + i, 1, month_data.get('month', ''))
            sheet.cell(row + 1 + i, 2, month_data.get('return', 0))
            sheet.cell(row + 1 + i, 3, month_data.get('trades', 0))
            sheet.cell(row + 1 + i, 4, month_data.get('win_rate', 0))

        self._apply_sheet_styles(sheet)

    def _apply_sheet_styles(self, sheet):
        """應用工作表樣式"""
        # 自動調整列寬
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # 最大寬度50
            sheet.column_dimensions[column_letter].width = adjusted_width

        # 應用邊框和對齊
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = self.styles['border']
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = self.styles['right']
                    else:
                        cell.alignment = self.styles['left']

    # 輔助方法
    def _evaluate_price(self, tech_data: Dict[str, Any]) -> str:
        """評價價格位置"""
        current = tech_data.get('current_price', 0)
        sma_20 = tech_data.get('sma_20', 0)

        if current > sma_20 * 1.05:
            return "高於均價"
        elif current < sma_20 * 0.95:
            return "低於均價"
        else:
            return "接近均價"

    def _evaluate_rsi(self, rsi: float) -> str:
        """評價RSI"""
        if rsi > 70:
            return "超買"
        elif rsi < 30:
            return "超賣"
        else:
            return "正常"

    def _evaluate_trend(self, trend_data: Dict[str, Any]) -> str:
        """評價趨勢"""
        trend = trend_data.get('trend', '')
        strength = trend_data.get('strength', 0)

        if '強勢' in trend:
            return "強勢"
        elif '多頭' in trend:
            return "多頭"
        elif '空頭' in trend:
            return "空頭"
        else:
            return "盤整"

    def _evaluate_position(self, position_data: Dict[str, Any]) -> str:
        """評價位階"""
        stage = position_data.get('stage', '')
        pct = position_data.get('current_pct', 50)

        if stage == '低檔':
            return "低估"
        elif stage == '高檔':
            return "高估"
        else:
            return "合理"

    def _evaluate_score(self, score_data: Dict[str, Any]) -> str:
        """評價綜合得分"""
        score = score_data.get('總分', 0)
        grade = score_data.get('等級', 'F')

        if grade in ['A+', 'A']:
            return "優秀"
        elif grade in ['B+', 'B']:
            return "良好"
        elif grade == 'C+':
            return "及格"
        else:
            return "待改善"

    def _calculate_ma_position(self, tech_data: Dict[str, Any], ma_type: str) -> str:
        """計算均線位置"""
        current = tech_data.get('current_price', 0)
        ma_value = tech_data.get(ma_type, 0)

        if ma_value == 0:
            return "無數據"

        ratio = (current - ma_value) / ma_value * 100
        if ratio > 5:
            return "上方"
        elif ratio < -5:
            return "下方"
        else:
            return "附近"

    def _get_rsi_status(self, rsi: float) -> str:
        """獲取RSI狀態"""
        if rsi > 70:
            return "超買區"
        elif rsi < 30:
            return "超賣區"
        else:
            return "中性區"

    def _get_macd_status(self, trend_data: Dict[str, Any]) -> str:
        """獲取MACD狀態"""
        macd_status = trend_data.get('macd_status', '')
        return macd_status if macd_status else "中性"